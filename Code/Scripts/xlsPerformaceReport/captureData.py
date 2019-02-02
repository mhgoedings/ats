from openpyxl import load_workbook

# import openpyxl
import csv
import os

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import sys

sys.path.append("/Users/szagar/ZTS/Dropbox/Business/ats/Code/Model/v3")
from base import Base, Session

from oos_test import OosTest

import pandas as pd


def parse_performance_summary(wb):
    sql_updates = {}
    ws = wb["Performance Summary"]

    # fields of interest
    foi = [
        "Total Net Profit",
        "Avg. Trade Net Profit",
        "Profit Factor",
        "Total Number of Trades",
        "Percent Profitable",
        "Avg. Winning Trade",
        "Avg. Losing Trade",
        "Sharpe Ratio",
        "Total Commission",
        "Total Slippage",
    ]
    db_map = {
        "Total Net Profit": "net_profit",
        "Avg. Trade Net Profit": "avg_trade",
        "Profit Factor": "profit_factor",
        "Total Number of Trades": "trade_count",
        "Percent Profitable": "percent_winners",
        "Avg. Winning Trade": "avg_winner",
        "Avg. Losing Trade": "avg_loser",
        "Sharpe Ratio": "sharpe_ratio",
        "Total Slippage": "slippage",
        "Total Commission": "commission",
        "Largest Winning Trade": "largest_winner",
        "Largest Losing Trade": "largest_loser",
        "Max. Consecutive Winning Trades": "win_streak",
        "Max. Consecutive Losing Trades": "lose_streak",
        "Avg. Bars in Total Trades": "avg_bars_in",
        "Avg. Bars in Winning Trades": "avg_bars_in_winner",
        "Avg. Bars in Losing Trades": "avg_bars_in_loser",
        "Avg. Bars in Even Trades": "avg_bars_in_even",
        "Max. Shares/Contracts Held": "max_contracts_held",
        "Total Shares/Contracts Held": "total_contracts_held",
        "Trading Period": "test_period_str",
        "Time in the Market": "time_in_str",
        "Max. Equity Run-up": "max_run_up",
        "Max. Drawdown (Intra-day Peak to Valley):Value": "drawdown_intra",
        "Max. Drawdown (Trade Close to Trade Close):Value": "drawdown_day",
        "Return on Account": "rtn_on_account",
        "RINA Index": "rina",
        "Max. Trade Drawdown": "drawdown_max",
        "Account Size Required": "required_account",
        "Return on Initial Capital": "rtn_on_init_cap",
        "Annual Rate of Return": "annual_ror",
        "Buy & Hold Return": "buy_hold_ror",
        "Avg. Monthly Return": "avg_monthly_rtn",
        "Std. Deviation of Monthly Return": "stdev_monthly_rtn",
        "Return Retracement Ratio": "rtn_retracement_ratio",
        "Max. Equity Run-up": "max_run_up",
        "Winning Trades": "win_count",
        "Losing Trades": "lose_count",
        "Even Trades": "even_count",
    }
    multiplier = {
        "percent_winners": 100,
        "buy_hold_ror": 100,
        "annual_ror": 100,
        "rtn_on_account": 100,
        "rtn_on_init_cap": 100,
    }
    prefix = ""
    for row in ws.rows:
        try:
            k = row[0].value  # .title
        except AttributeError:
            continue
        if k in (None, " ", ""):
            prefix = ""
            continue
        k = f"{prefix}{k}"
        if k == "All Trades":
            break
        if row[1].value == None:
            prefix = k + ":"

        # if k in foi:

        print(f"if {k} in {db_map.keys()}:")
        if k in db_map.keys():
            sql_updates[db_map[k]] = row[1].value  # * multiplier.get(db_map[k],1.0)
            if db_map[k] in multiplier:
                sql_updates[db_map[k]] = sql_updates[db_map[k]] * multiplier[db_map[k]]
        else:
            # print(f'{k} NOT in foi')
            pass

    return sql_updates


def parse_trades_list(tn, wb):
    ws = wb["Trades List"]
    fn = f"test_{tn}"
    print(f"write trades to {msa_dir}{fn}.csv")
    with open(msa_dir + fn + ".csv", "w") as f:
        c = csv.writer(f)
        for row in ws.rows:
            # print([cell.value for cell in row])
            c.writerow([cell.value for cell in row])

    wb.close()


def parse_monthly(wb):
    ws = wb["Monthly"]
    mdata = False
    mdata_cols = False
    data = []
    data_cols = []
    for row in ws.rows:
        try:
            k = row[0].value  # .title
        except AttributeError:
            continue
        if k is None:
            continue
        if k == "Mark-To-Market Rolling Period Analysis:":
            break
        if mdata:
            data.append([str(cell.value) for cell in row])
        if mdata_cols:
            mdata_cols = False
            mdata = True
            for r in row:
                data_cols.append(r.value)
        if k == "Mark-To-Market Period Analysis:":
            mdata_cols = True

    # print(f'data_cols={data_cols}')
    # for d in data:
    #    print(','.join(d))
    df = pd.DataFrame(data, columns=data_cols)
    return df


def writeTrades2csv(tn, wb_file):
    wb = load_workbook(filename=wb_file)

    sheet = wb["Trades List"]
    fn = f"test_{tn}"
    with open(msa_dir + fn + ".csv", "w") as f:
        c = csv.writer(f)
        for row in sheet.rows:
            # print([cell.value for cell in row])
            c.writerow([cell.value for cell in row])


def xlsMetrics2db(test, wb):
    db_updates = parse_performance_summary(wb)
    print(db_updates)
    try:
        session.query(OosTest).filter(OosTest.id == test.id).update(db_updates)
        session.commit()
    except:
        session.rollback()
        raise
    set_status(test, "captured")


def set_status(test, status_state):
    updates = {"status": "OosTest", "status_state": status_state}
    try:
        session.query(OosTest).filter(OosTest.id == test.id).update(updates)
        session.commit()
    except:
        session.rollback()
        raise


def all_query():
    return session.query(OosTest)

def id_query(id):
    return session.query(OosTest).filter(OosTest.id == id)


def test_query():
    return session.query(OosTest).filter(
        OosTest.status == "OosTest", OosTest.status_state == "done"
    )


##########################################################

dbox = "/Users/szagar/ZTS/Dropbox/"
msa_dir = dbox + "Business/ats/Data/MSA/StrategyTrades/"
ws_dir = f"{dbox}Business/ats/Data/Queue/OosResults"

db = Session()
db_url = os.environ["DB_HEROKU"]

engine = create_engine(db_url)
Base.metadata.bind = engine

DBSession = sessionmaker(bind=engine)
session = DBSession()


#for test in id_query(41):
#for test in all_query():
for test in test_query():
    file = f"{ws_dir}/{test.id}.xlsx"
    print(f"test {test.id}  {test.symbol}  {test.oos_file}  perf_file: {file}")

    wb = load_workbook(filename=file, read_only=True)
    df = parse_monthly(wb)
    df.set_index("Period", inplace=True)
    df[["Net Profit", "Profit Factor"]] = df[["Net Profit", "Profit Factor"]].apply(
        pd.to_numeric
    )

    stock_return = df["Net Profit"]
    # stock_return.apply(lambda x: x / x[0])
    stock_return = df["Net Profit"] / df["Net Profit"][0]
    # print(stock_return.head())
    # print(df)

    xlsMetrics2db(test, wb)

    print(f"write trades to {msa_dir}")
    parse_trades_list(test.id, wb)
    wb.close()

    # break
exit()

"""
stock_return = df['Net Profit']
#stock_return.apply(lambda x: x / x[0])
stock_return = df['Net Profit'] / df['Net Profit'][0]
stock_return.head()
"""
