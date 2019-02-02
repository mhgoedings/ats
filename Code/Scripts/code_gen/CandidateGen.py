from jinja2 import FileSystemLoader, Environment
from openpyxl import Workbook
from openpyxl import load_workbook
import re

dp = {}
desc ={}
inputs = {}
varz = {}
InputsVars = ['longShort', 'sessionStart', 'sessionEnd', 
            'tradesPerDay', 'dayOrSwing', 'poiSwitch',
            'natr', 'fract', 
            'filter1switch', 'filter1n1', 'filter1n2',
            'filter2switch', 'filter2n1', 'filter2n2',
             'tSegment', 'slSwitch', 'stopLoss',
             'ptSwitch', 'profitTarget',
            ]

def parseExcelRow(row):
    dpTemplate = row[1].value
    dp['dataSet'] = row[2].value
    dp['dataBlock'] = row[3].value
    dp['symbol'] = row[8].value
    dp['timeFrame1'] = row[9].value
    dp['timeFrame1unit'] = row[10].value
    dp['use2ndDataSet'] = row[19].value
    if dp['use2ndDataSet'] == 'TRUE':
        dp['timeFrame2'] = row[11].value
        dp['timeFrame2unit'] = row[12].value
    dp['fitnessFunc'] = row[13].value
    dp['maxDaysBack'] = row[14].value
    dp['sessionNum'] = row[15].value
    dp['startDt'] = row[17].value.strftime("%m/%d/%y")
    dp['endDt'] = row[18].value.strftime("%m/%d/%y")
    dp['longShort'] = row[20].value
    dp['sessionStart'] = row[21].value
    dp['sessionEnd'] = row[22].value
    dp['tradesPerDay'] = row[23].value
    dp['dayOrSwing'] = row[24].value
    dp['poiSwitch'] = row[25].value
    dp['natr'] = row[27].value
    dp['fract'] = row[28].value
    dp['filter1switch'] = row[29].value
    dp['filter1n1'] = row[30].value
    dp['filter1n2'] = row[31].value
    dp['filter2switch'] = row[32].value
    dp['filter2n1'] = row[33].value
    dp['filter2n2'] = row[34].value
    dp['tSegment'] = row[35].value
    dp['slSwitch'] = row[36].value
    dp['stopLoss'] = row[37].value
    dp['ptSwitch'] = row[38].value
    dp['profitTarget'] = row[39].value
    
    
def readDbRecord(dpNumber):
    wb = load_workbook(filename = 'atsPrototypes.xlsx')
    wb.close()
    ws = wb['Proto Runs']
    start, stop = 1, 100    # This will allow you to set a lower and upper limit
    for index, row in enumerate(ws.iter_rows()):
        if start < index < stop:
            #print(f'index={index}  row[0]={row[0].value}')
            if row[0].value == dpNumber:
                return row
                #print([v.value for v in row])



def defineInputsVars():
    #pattern = re.compile(r'^\w+,\w+,\w+$')
    pattern = re.compile(r'^.+,.+,.+$')
    for i in InputsVars:
        if i in dp:
            match = re.search(pattern, str(dp[i]))
            if match:
                if i not in inputs:
                    inputs[i] = {}
                inputs[i]['optStr'] = dp[i]
                inputs[i]['baseVal'] = 'tbd' #candData[i]
                inputs[i]['dtype'] = 'tbd'   #varRef[i]
            else:
                if i not in varz:
                    varz[i] = {}
                varz[i]['baseVal'] = dp[i] #candData[i]
                varz[i]['dtype'] = 'tbd'   #varRef[i]
                #print(f'{i}({dp[i]}) :Not found')


def render_from_template(directory, template_name, **kwargs):
    loader = FileSystemLoader(directory)
    env = Environment(loader=loader)
    template = env.get_template(template_name)
    return template.render(**kwargs)

def formatCommentStrings():
    desc['chart_setup'] = f'''Chart Setup:
        symbol = {dp['symbol']}
        1st timeframe = {dp['timeFrame1']} {dp['timeFrame1unit']}
        Fitness Function = {dp['fitnessFunc']}
        maxDaysBack = {dp['maxDaysBack']}
        session number = {dp['sessionNum']}
        start date = {dp['startDt']}
        end date = {dp['endDt']}'''
    desc['prototype_info'] = f'''Prototype Info::
        dataSet = {dp['dataSet']}
        dataBlock = {dp['dataBlock']}    
        '''

def processTemplate():
    templateLoader = FileSystemLoader(searchpath="./")
    templateEnv = Environment(loader=templateLoader)
    TEMPLATE_FILE = 'cand_template.txt'
    template = templateEnv.get_template(TEMPLATE_FILE)
    outputText = template.render(chart_setup=desc['chart_setup'],
                                 prototype_info=desc['prototype_info'],
                                 dp_data=dp,
                                 comments=desc,
                                 inputs = inputs,
                                 varz = varz,
                                )
    #print(outputText)
    return(outputText)


def generateEL(candData):
    dpNumber = candData['dpNumber']
    row = readDbRecord(dpNumber)
    parseExcelRow(row)
    defineInputsVars()
    formatCommentStrings()
    print(f'Inputs: {inputs}')
    #print(f'Vars: {varz}')
    return processTemplate()
    
out = generateEL({'dpNumber': 21})
print(out)
